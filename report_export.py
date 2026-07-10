"""
report_export.py — render report data structures to PDF or Excel.

A report is a plain dict so any route can build one without importing render libs:

    {
      'title': 'Financial Statements',
      'subtitle': '2026-01-01 to 2026-12-31',
      'sections': [
        {'heading': 'Income Statement',
         'columns': ['', 'Amount'],          # optional header row
         'rows': [
            {'cells': ['Loan interest earned', 9000.0]},
            {'cells': ['Net surplus', -13000.0], 'bold': True},
         ]},
      ],
    }

Numeric cells (int/float) are formatted as currency and right-aligned; strings
are left-aligned. PDF uses reportlab (always available); Excel uses openpyxl
(optional — the route degrades gracefully if it isn't installed).
"""

import io
from datetime import datetime

from flask import make_response, flash, redirect, request


def _numeric_columns(section):
    """Column indices whose data cells are all numeric."""
    rows = section.get('rows', [])
    if not rows:
        return set()
    width = max(len(r['cells']) for r in rows)
    numeric = set()
    for col in range(width):
        vals = [r['cells'][col] for r in rows if col < len(r['cells'])]
        if vals and all(isinstance(v, (int, float)) for v in vals):
            numeric.add(col)
    return numeric


def _fmt(cell):
    return f"₦{cell:,.2f}" if isinstance(cell, (int, float)) else str(cell)


def to_pdf(report):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=report['title'],
                            topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    h1  = ParagraphStyle('h1', parent=styles['Title'], fontSize=16, spaceAfter=2)
    sub = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9, textColor=colors.grey)
    hs  = ParagraphStyle('hs', parent=styles['Heading2'], fontSize=12, spaceBefore=10, spaceAfter=4)

    elems = [Paragraph(report['title'], h1)]
    if report.get('subtitle'):
        elems.append(Paragraph(report['subtitle'], sub))
    elems.append(Spacer(1, 8))

    for sec in report['sections']:
        if sec.get('heading'):
            elems.append(Paragraph(sec['heading'], hs))
        numeric = _numeric_columns(sec)
        data, cols = [], sec.get('columns')
        if cols:
            data.append([str(c) for c in cols])
        header_offset = 1 if cols else 0
        bold_rows = []
        for i, r in enumerate(sec['rows']):
            data.append([_fmt(c) for c in r['cells']])
            if r.get('bold'):
                bold_rows.append(i + header_offset)
        if not data:
            continue

        table = Table(data, hAlign='LEFT')
        st = [('FONTSIZE', (0, 0), (-1, -1), 9),
              ('TOPPADDING', (0, 0), (-1, -1), 3),
              ('BOTTOMPADDING', (0, 0), (-1, -1), 3)]
        for col in numeric:
            st.append(('ALIGN', (col, 0), (col, -1), 'RIGHT'))
        if cols:
            st += [('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                   ('BACKGROUND', (0, 0), (-1, 0), colors.whitesmoke),
                   ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.grey)]
        for br in bold_rows:
            st.append(('FONTNAME', (0, br), (-1, br), 'Helvetica-Bold'))
            st.append(('LINEABOVE', (0, br), (-1, br), 0.4, colors.lightgrey))
        table.setStyle(TableStyle(st))
        elems.append(table)
        elems.append(Spacer(1, 6))

    elems.append(Spacer(1, 10))
    elems.append(Paragraph(f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}", sub))
    doc.build(elems)
    buf.seek(0)
    return buf.getvalue()


def to_xlsx(report):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    grey = Font(color='888888', italic=True)
    hdr_fill = PatternFill('solid', fgColor='F0F0F0')

    row = 1
    ws.cell(row, 1, report['title']).font = Font(size=14, bold=True)
    row += 2
    if report.get('subtitle'):
        ws.cell(row, 1, report['subtitle']).font = grey
        row += 2

    for sec in report['sections']:
        if sec.get('heading'):
            ws.cell(row, 1, sec['heading']).font = Font(size=12, bold=True)
            row += 1
        cols = sec.get('columns')
        if cols:
            for ci, h in enumerate(cols, 1):
                c = ws.cell(row, ci, str(h))
                c.font = Font(bold=True)
                c.fill = hdr_fill
            row += 1
        for r in sec['rows']:
            for ci, cell in enumerate(r['cells'], 1):
                c = ws.cell(row, ci)
                if isinstance(cell, (int, float)):
                    c.value = float(cell)
                    c.number_format = '#,##0.00'
                    c.alignment = Alignment(horizontal='right')
                else:
                    c.value = str(cell)
                if r.get('bold'):
                    c.font = Font(bold=True)
            row += 1
        row += 1

    for col_cells in ws.columns:
        width = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(width + 2, 12), 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def report_response(report, fmt, redirect_url=None):
    """Return a downloadable PDF or Excel response for a report dict.
    Falls back to a flash + redirect if Excel is requested but openpyxl is absent."""
    fmt = (fmt or 'pdf').lower()
    slug = ''.join(ch if ch.isalnum() else '_' for ch in report['title'].lower()).strip('_')

    if fmt in ('xlsx', 'excel'):
        try:
            data = to_xlsx(report)
        except ImportError:
            flash('Excel export needs the openpyxl package. Use PDF, or install openpyxl.', 'warning')
            return redirect(redirect_url or request.path)
        mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        ext = 'xlsx'
    else:
        data = to_pdf(report)
        mime = 'application/pdf'
        ext = 'pdf'

    resp = make_response(data)
    resp.headers['Content-Type'] = mime
    resp.headers['Content-Disposition'] = f'attachment; filename={slug}.{ext}'
    return resp
